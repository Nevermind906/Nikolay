# coding=utf-8
import utility
import exceptions
from openpyxl import load_workbook
from datetime import datetime, time, date, timezone, timedelta

class ExcelIO:
    def __init__(self):
        self.filename = './data.xlsx'
        self.wb = load_workbook(self.filename)
        self.schedule = self.wb['Расписание']
        self.subjects = self.wb['Предметы']
        self.info = self.wb['Информация']

    def __get_last_row(sheet):
        i = 1
        while sheet.cell(row=i, column=1).value:
            i = i + 1
        return sheet[i]

    def __get_schedule_cell(self, date, num):
        i = 2
        while self.schedule.cell(row=i, column=1).value:
            if self.schedule.cell(row=i, column=1).value.date() == date:
                return self.schedule.cell(row=i, column=num+1)
            i = i + 1
        raise exceptions.NikolayException("Не удалось найти указанную дату в расписании.")

    def get_schedule(self, date, delta=1):
        response = ""
        i = 2
        while self.schedule.cell(row=i, column=1).value:
            if self.schedule.cell(row=i, column=1).value.date() == date:
                for k in range(i, i+delta):
                    day = date + timedelta(k - i)
                    response = response + utility.weekdays[day.weekday()] + " " + str(day) + "\n"
                    daily = ""
                    for j in range(2, 9):
                        cl = self.schedule.cell(row=k, column=j)
                        if cl.value:
                            daily = daily + str(j-1) + ") " + cl.value + '\n'
                    if daily == "":
                        daily="Выходной, отдыхаем\n\n"
                    response = response + daily + '\n'
                return response
            i = i + 1
        raise exceptions.NikolayException("Не удалось найти указанную дату в расписании.")

    def schedule_read(self, date, num):
        return self.__get_schedule_cell(date, num).value

    def schedule_write(self, date, num, str):
        cell = self.__get_schedule_cell(date, num)
        cell.value = str

    def get_info(self, q, date):
        i = 2
        count = 0
        messages = []
        while self.info.cell(row=i, column=1).value:
            added = self.info.cell(row=i, column=1).value
            deadline = self.info.cell(row=i, column=2).value
            text = self.info.cell(row=i, column=3).value
            attachment = self.info.cell(row=i, column=4).value
            if deadline.date() >= date and q in text:
                count = count + 1
                message = "%d. дата добавления: %s\nдата окончания: %s\n" \
                            % (i, added.date(), deadline.date()) + text
                messages.append((message, attachment))
            i = i + 1
        if not count:
            raise exceptions.NikolayException('Информация отсутствует')
        return messages

    def info_read(self, id):
        return [id] + [cell.value for cell in self.info[id]]

    def info_write(self, added, deadline, message, attachment, id=0):
        if not id:
            row =  ExcelIO.__get_last_row(self.info)
        else:
            row = self.info[id]
        row[0].value = added
        row[1].value = deadline
        row[2].value = message
        row[3].value = attachment

    def info_delete(self, id):
        if self.info[id][0].value:
            self.info.delete_rows(id)
        else:
            raise exceptions.NikolayException('Не удалось найти сообщение с id=%d' % id)

    def get_subjects(self):
        response = []
        i = 2
        while self.subjects.cell(row=i, column=1).value:
            row_val = [cl.value for cl in self.subjects[i] if cl]
            response.append(", ".join(row_val))
            i = i + 1
        return sorted(response)

    def save(self):
        self.wb.save(filename=self.filename)
